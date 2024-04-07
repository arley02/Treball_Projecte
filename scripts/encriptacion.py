import bcrypt
import openpyxl
from os import system

# Funció per crear una contrasenya xifrada
def create_hash(password):
    salt = bcrypt.gensalt()
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed_password

# Funció per comprovar la contrasenya
def check_password(input_password, hashed_password):
    return bcrypt.checkpw(input_password.encode('utf-8'), hashed_password)

# Carrega de noms d'usuari i contrasenyes xifrades des d'un fitxer Excel
def load_credentials(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    credentials = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        username = row[0]
        hashed_password = row[1].encode('utf-8')  # Convertir la contraseña a bytes
        credentials[username] = hashed_password
    return credentials

# Afegir un nou compte
def datos_encript(file_path, username, password):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.append((username, create_hash(password)))
    wb.save(file_path)

# Accedir al sistema

# Ruta al fitxer Excel amb noms d'usuari i contrasenyes
file_path = "PROG/credentials.xlsx"

# Carregar noms d'usuari i contrasenyes xifrades
credentials = load_credentials(file_path)