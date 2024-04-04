import bcrypt
import openpyxl
from os import system

# Funció per crear una contrasenya xifrada
def create_hash(password):
    salt = bcrypt.gensalt()
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed_password

# Funció per comprovar la contrasenya
def check_password(input_password, hashed_password):
    return bcrypt.checkpw(input_password.encode('utf-8'), hashed_password)

# Carrega de noms d'usuari i contrasenyes xifrades des d'un fitxer Excel
def load_credentials(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    credentials = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        username = row[0]
        hashed_password = row[1]
        credentials[username] = hashed_password
    return credentials

# Afegir un nou compte
def add_account(file_path, username, password):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.append((username, create_hash(password)))
    wb.save(file_path)

# Accedir al sistema
def login(credentials):
    username = input("Introduïu el nom d'usuari: ")
    password = input("Introduïu la contrasenya: ")

    if username in credentials and check_password(password, credentials[username]):
        print("S'ha iniciat sessió amb èxit!")
        # Aquí podeu afegir el codi per a realitzar l'acció X
    else:
        print("Nom d'usuari o contrasenya incorrectes!")

# Ruta al fitxer Excel amb noms d'usuari i contrasenyes
file_path = "PROG/credentials.xlsx"

# Carregar noms d'usuari i contrasenyes xifrades
credentials = load_credentials(file_path)

# Selecció del mode: iniciar sessió o crear un nou compte
mode = input("Seleccioneu el mode:\n 1 - Iniciar sessió\n 2 - Crear un compte\n \n Seleccione una opcio: ")

if mode == "1":
    system("cls")
    login(credentials)
elif mode == "2":
    system("cls")
    new_username = input("Introduïu el nou nom d'usuari: ")
    new_password = input("Introduïu la nova contrasenya: ")
    add_account(file_path, new_username, new_password)
    print("El compte s'ha creat amb èxit!")
else:
    print("Mode incorrecte. Si us plau, seleccioneu 1 o 2.")
